import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

def generate_excel_report(active_rule_ids, all_rules_config, output_path):
    """Generates an Excel report of the AWS Config rules."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AWS Config Rules"

    # Headers
    headers = ["Rule ID", "Rule Name", "Description", "Rationale", "Remediation"]
    ws.append(headers)

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data
    for rule_id in sorted(active_rule_ids):
        if rule_id in all_rules_config:
            rule = all_rules_config[rule_id]
            ws.append([
                rule_id,
                rule.get('name', ''),
                rule.get('description', ''),
                rule.get('rationale', ''),
                rule.get('remediation', '')
            ])

    # Column widths and cell styling
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 80
    ws.column_dimensions['E'].width = 80

    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=2, max_col=5, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border

    wb.save(output_path)
